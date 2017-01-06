
# coding: utf-8

# In[1]:

#builds invoice summary using 'Invoice Summary - Large Template.xlsx' and a given Media Ops Campaign Line Item Report in .csv format from a campaign's SF opportunity
#data structure: summary (dictionary) / component (list) / placement (dictionary)
#separate list of components ordered chronologically is used to process components in the summary in order
#separate list of placements in each component ordered alphabetically is used to process each component's placements in order

import sys
import csv
import re
import itertools
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from decimal import Decimal, ROUND_HALF_UP

#prompts the user to either enter the file name of an IOC in .csv format located in the same folder as the program file or to quit the program

filename = input('Enter file name or enter "Q" to quit this program: \n\n')

if str.lower(filename) == 'q':
    print('\n Program terminated.')
    sys.exit()

else: pass


while True:
    if not filename.endswith('csv'):
        filename = input('Please enter a .csv file name: \n\n')
        continue

    else:
        break

while True:
    try:
        IOC = csv.DictReader(open(filename))
        print('\n\n Now reading through', filename, '... \n\n')
        break

    except:
        filename = input('Please enter a valid file name: \n\n')
        continue

#function to determine a placement's category based on the placement name

def categorization(target_name):

    if re.search('URL', target_name):
        category = 'URL'

    elif re.search('NL', target_name) or re.search('Newsletter', target_name, flags=re.IGNORECASE):
        category = 'Newsletter'

    elif re.search('CAE', target_name, flags=re.IGNORECASE) or re.search('Exclusivity', target_name, flags=re.IGNORECASE):
        category = 'CAE'
        
    elif re.search('AE', target_name) or re.search('Extend', target_name, flags=re.IGNORECASE) or re.search('Extension', target_name, flags=re.IGNORECASE):
        category = 'Turn'

    elif re.search('TOC', target_name, flags=re.IGNORECASE) or re.search('Roadblock', target_name, flags=re.IGNORECASE):
        category = 'Roadblock'

    elif re.search('Showcase', target_name, flags=re.IGNORECASE) or re.search('Build Fee', target_name, flags=re.IGNORECASE) or re.search('IMU', target_name) or re.search('CSU', target_name):
        category = 'IMU'

    elif re.search('Medpulse', target_name, flags=re.IGNORECASE):
        category = 'Medpulse'

    elif re.search('Video', target_name, flags=re.IGNORECASE) or re.search('Preroll', target_name, flags=re.IGNORECASE):
        category = 'Video'

    elif re.search('Gateway', target_name, flags=re.IGNORECASE):
        category = 'Gateway'

    elif re.search('Takeover', target_name, flags=re.IGNORECASE):
        category = 'Channel Takeover'

    elif re.search('Conference', target_name, flags=re.IGNORECASE):
        category = 'Conference Package'

    elif re.search('Select', target_name, flags=re.IGNORECASE) or re.search('First Imp', target_name, flags=re.IGNORECASE):
        category = 'Other'

    elif re.search('M01', target_name, flags=re.IGNORECASE) or re.search('320x50', target_name, flags=re.IGNORECASE) or re.search('App', target_name) or re.search('300x50', target_name, flags=re.IGNORECASE):
        category = 'Mobile'

    else:
        category = 'Media'

    return category

#allows the program to comprehend the date format used in the IOC

def date_format(date):

    formatted_date = datetime.strptime(date, "%m/%d/%Y")
    return formatted_date

#asks whether or not to include reservation notes and/or targeting details in the invoice summary

include_notes = ''
notes_or_targeting = ''

include_notes = input('Include reservation notes and/or targeting details in invoice summary? Y/N? \n\n')

while True:
    if str.lower(include_notes) != 'y' and str.lower(include_notes) != 'n':
        include_notes = input('\n\n Please enter either "Y" or "N" only. \n\n')
        continue

    else:
        break

if str.lower(include_notes) == 'y':

    notes_or_targeting = input('Include (R)eservation notes or (T)argeting details? R/T? \n\n')
        
    while True:
        if str.lower(notes_or_targeting) != 'r' and str.lower(notes_or_targeting) != 't':
            notes_or_targeting = input('\n\n Please enter either "R" or "T" only. \n\n')
            continue
        
        else:
            break
    
#reads each row in order to build a list of components based on start/end dates and category (with or without reservation notes/targeting details included)
#ignores the text below the media plan and skips placements with 0 impression goal and $0 cost

componentlist = list()

for placement in IOC:
    if re.search('[a-z]', placement['Opportunity Number'], flags=re.IGNORECASE):
        continue

    elif placement['Opportunity Number'] == '':
        continue

    elif placement['Approved Reservation Amount'] == '0' and placement['Imps to Reserve'] == '0':
        continue
        
    else:
        if notes_or_targeting == 'r':
            component = (placement['Reservation Start Date'], placement['Reservation End Date'], categorization(placement['Inventory Target Record: Inventory Target Name']), placement['Reservation Notes'])
        elif notes_or_targeting == 't':
            component = (placement['Reservation Start Date'], placement['Reservation End Date'], categorization(placement['Inventory Target Record: Inventory Target Name']), placement['Targeting Details'])
        elif include_notes == 'n':
            component = (placement['Reservation Start Date'], placement['Reservation End Date'], categorization(placement['Inventory Target Record: Inventory Target Name']), '')
        if component not in componentlist:
            componentlist.append(component)
        else: continue

#orders the component list based on 1) start date, 2) end date, 3) category and adds the components to summary

summary = dict()

ordered_componentlist = sorted(componentlist, key = lambda component: (date_format(component[0]), date_format(component[1]), component[2]))

for entry in ordered_componentlist:
    summary[entry] = list()

#sorts each placement into the appropriate component list based on its start/end dates, category, and reservation note/targeting details
#ignores the text below the media plan and skips placements with 0 impression goal and $0 cost
#keeps track of how many placements were skipped

IOC = csv.DictReader(open(filename))

skipped = 0

for placement in IOC:
    if placement['Approved Reservation Amount'] == '0' and placement['Imps to Reserve'] == '0':
        print('SKIPPED:', placement['Inventory Target Record: Inventory Target Name'], placement['Imps to Reserve'], 'Impressions', placement['Approved Unit Price'], 'CPM', placement['Approved Reservation Amount'], 'Placement Cost')
        skipped = skipped + 1
        continue
    elif re.search('[a-z]', placement['Opportunity Number'], flags=re.IGNORECASE):
        continue

    elif placement['Opportunity Number'] == '':
        continue

    else:
        if notes_or_targeting == 'r':
            component = (placement['Reservation Start Date'], placement['Reservation End Date'], categorization(placement['Inventory Target Record: Inventory Target Name']), placement['Reservation Notes'])
        elif notes_or_targeting =='t':
            component = (placement['Reservation Start Date'], placement['Reservation End Date'], categorization(placement['Inventory Target Record: Inventory Target Name']), placement['Targeting Details'])
        elif include_notes == 'n':
            component = (placement['Reservation Start Date'], placement['Reservation End Date'], categorization(placement['Inventory Target Record: Inventory Target Name']), '')
        for key in summary:
            if key == component:
                summary[key].append(placement)

#orders the placements in each component based on placement names

for component in summary:
    ordered_placementlist = sorted(summary[component], key = lambda placement: placement['Inventory Target Record: Inventory Target Name'])
    summary[component] = ordered_placementlist

#writes the data to the invoice summary template

print('\n Now writing', filename, 'to invoice summary... \n\n')

template = load_workbook('Invoice Summary - Large Template.xlsx')
invoice_summary = template.active

#fills in the 'SF', 'Start Date', 'End Date', and 'Total Cost' fields at the top of the invoice summary

SF_num = 0
placement_costs = list()
start_dates = list()
end_dates = list()

for component, placements in summary.items():
    for placement in placements:
        SF_num = int(placement['Opportunity Number'])
        placement_costs.append(float(placement['Approved Reservation Amount']))
        start_dates.append(placement['Reservation Start Date'])
        end_dates.append(placement['Reservation End Date'])

ordered_start_dates = sorted(start_dates, key = lambda date: date_format(date))
ordered_end_dates = sorted(end_dates, key = lambda date: date_format(date), reverse = True)

invoice_summary['B2'] = SF_num
invoice_summary['B6'] = sum(placement_costs)
invoice_summary['B7'] = ordered_start_dates[0]
invoice_summary['B8'] = ordered_end_dates[0]

#finds the row numbers associated with the 'Sequence Number' cells belonging to each component on the invoice summary template

sequence_locations = list()
a = len(ordered_componentlist)

for line in invoice_summary['A']:
    if a == 0: break
    if line.value == 'Sequence Number':
        sequence_locations.append(line.row)
        a = a - 1

#determines the ad size for a given placement based on its name

def get_ad_size(target_name):

    if re.search('Video', target_name, flags=re.IGNORECASE) or re.search('Preroll', target_name, flags=re.IGNORECASE):
        ad_size = 'Video'

    elif re.search('M01', target_name, flags=re.IGNORECASE) or re.search('320x50', target_name, flags=re.IGNORECASE) or re.search('300x50', target_name, flags=re.IGNORECASE):
        ad_size = '320x50'

    elif re.search('160x600', target_name, flags=re.IGNORECASE):
        ad_size = '160x600'

    elif re.search('300x250', target_name, flags=re.IGNORECASE):
        ad_size = '300x250'

    elif re.search('300x600', target_name, flags=re.IGNORECASE):
        ad_size = '300x600'

    elif re.search('728x90', target_name, flags=re.IGNORECASE):
        ad_size = '728x90'

    else:
        ad_size = 'All'

    return ad_size

#creates the ordered list of months in the cells above the 'Delivered' cells in each component starting with the month the campaign begins in

start_date = date_format(ordered_start_dates[0])
calendar = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 0: 'December'}
ordered_number_months = list()

for month in range(start_date.month, start_date.month + 17):
    ordered_number_months.append(month % 12)

written_months = list()
    
for month in ordered_number_months:
    written_month = calendar[month]
    written_months.append(written_month)

months_cycle = itertools.cycle(written_months)
ordered_written_months = list()

b = 0
for month in itertools.cycle(months_cycle):
    ordered_written_months.append(month)
    b = b + 1
    if b >= 17: break
        
#fills in each component's headers and month columns with correct style formatting
#fills in each component's placements and corresponding fields
#hides unused rows in each component
#appends 'Bonus' to bonus placement names
#highlights placements where the 'Approved Reservation Amount' field doesn't match the calculated cost or the 'Imps to Reserve' field is blank in red
#replaces each red highlighted row's 'Total Cost' formula cell with that placement's 'Approved Reservation Amount' and omits the 'Approved Unit Price'
#keeps track of the number of red highlighted placements in the invoice summary

c = 0
d = 14
e = 0
f = 4

highlighted = 0

for component, sequence in zip(ordered_componentlist, sequence_locations):

    invoice_summary.cell(row = sequence + 1, column = 1).value = component[0]
    invoice_summary.cell(row = sequence + 1, column = 2).value = component[1]
    invoice_summary.cell(row = sequence + 1, column = 3).value = component[0]
    invoice_summary.cell(row = sequence + 2, column = 2).value = component[2]
    invoice_summary.cell(row = sequence + 2, column = 4).value = component[3]
    invoice_summary.cell(row = sequence + 2, column = 2).font = Font(name = 'Times New Roman', size = 14, bold = True)
    invoice_summary.cell(row = sequence + 2, column = 2).alignment = Alignment(horizontal='center')
    invoice_summary.cell(row = sequence + 2, column = 2).fill = PatternFill(fill_type = 'solid', start_color = 'C0C0C0', end_color = 'C0C0C0')

    while c < 17:
        invoice_summary.cell(row = sequence + 2, column = d).value = ordered_written_months[c]
        d = d + 6
        c = c + 1
     
    c = 0
    d = 14
    
    for row in range(sequence + 4 + len(summary[component]), sequence + 45, 1):
        invoice_summary.row_dimensions[row].hidden = True

    for placement in summary[component]:
        
        calc_cost = (Decimal(placement['Approved Unit Price'])*Decimal(placement['Imps to Reserve']))/Decimal(1000)
        
        if float(placement['Approved Unit Price']) == 0 and float(placement['Approved Reservation Amount']) == 0:
            invoice_summary.cell(row = sequence + f, column = 1).value = 'Bonus ' + summary[component][e]['Inventory Target Record: Inventory Target Name']
        else:
            invoice_summary.cell(row = sequence + f, column = 1).value = summary[component][e]['Inventory Target Record: Inventory Target Name']
        
        if int(placement['Imps to Reserve']) == 1 or Decimal(calc_cost).quantize(Decimal('0.01'), ROUND_HALF_UP) != Decimal(placement['Approved Reservation Amount']) or placement['Imps to Reserve'] == '':
            invoice_summary.cell(row = sequence + f, column = 1).fill = PatternFill(fill_type = 'solid', start_color = 'ff0000', end_color = 'ff0000')
            invoice_summary.cell(row = sequence + f, column = 7).fill = PatternFill(fill_type = 'solid', start_color = 'ff0000', end_color = 'ff0000')
            invoice_summary.cell(row = sequence + f, column = 10).fill = PatternFill(fill_type = 'solid', start_color = 'ff0000', end_color = 'ff0000')
            invoice_summary.cell(row = sequence + f, column = 7).value = ''
            invoice_summary.cell(row = sequence + f, column = 10).value = float(placement['Approved Reservation Amount'])
            
            print('HIGHLIGHTED:', placement['Inventory Target Record: Inventory Target Name'], placement['Approved Reservation Amount'], 'Placement Cost', float(calc_cost), 'Calculated Cost')
            highlighted = highlighted + 1
        
        else:
            invoice_summary.cell(row = sequence + f, column = 7).value = float(summary[component][e]['Approved Unit Price'])
            
        invoice_summary.cell(row = sequence + f, column = 2).value = get_ad_size(summary[component][e]['Inventory Target Record: Inventory Target Name'])
        invoice_summary.cell(row = sequence + f, column = 3).value = int(summary[component][e]['Imps to Reserve'])
        
        e = e + 1
        f = f + 1
        
    e = 0
    f = 4


    print('\n', component, 'finished. \n')

#finds the row number associated with grand total row at the bottom of the invoice summary

grand_total = 0

for line in invoice_summary['A']:
    if line.value == 'Grand Total':
        grand_total = line.row
        break

#deletes the 'Sequence Number' cells belonging to each of the unused components in the invoice summary in order to exclude them when uploading to Billing Tracker

for line in invoice_summary['A']:
    if line.value == 'Sequence Number' and line.row not in sequence_locations:
        line.value = ''
        
#hides all of the unused components between the last used component and the grand total row

for row in range(sequence_locations[-1] + 46, grand_total - 2, 1):
    invoice_summary.row_dimensions[row].hidden = True

#reports how many placements were skipped and highlighted

print('\n', skipped, 'placements skipped due to having a goal of 0 and a CPM of $0.00.')
print('\n', highlighted, 'placements highlighted in red on the invoice summary for having an approved reservation amount that does not match the calculated cost.\n')
print('Each of these red highlighted placements have had its total cost formula cell replaced with its approved reservation amount and its approved unit price omitted.')

#attempts to save the invoice summary to a .xlsx file and prompts the user to try to save the file again or quit the program if the save attempt fails

save_name = 'SF' + str(SF_num) + ' Invoice Summary.xlsx'
        
while True:
    try:
        template.save(save_name)
        print('\nThe invoice summary for', 'SF'+str(SF_num), 'is done.')
        quit = input('\nEnter any key to quit.')
        break
    
    except:
        while True:
            print('\nInvoice summary failed to save.', save_name, 'must be closed before the file can be saved. \n')
            save_or_quit = input('\nPress "S" to try to save the file again or press "Q" to quit this program. \n' )
                
            if str.lower(save_or_quit) == 's': break
            
            elif str.lower(save_or_quit) == 'q':
                print('\n Program terminated.')
                sys.exit()
        
            else:
                while True:
                    save_or_quit = input('\n\nPlease enter either "S" or "Q" only. \n\n')
                    if str.lower(save_or_quit) == 's': break
                    
                    elif str.lower(save_or_quit) == 'q':
                        print('\n Program terminated.')
                        sys.exit()
                    
                    else: continue
            
            break


# In[ ]:



